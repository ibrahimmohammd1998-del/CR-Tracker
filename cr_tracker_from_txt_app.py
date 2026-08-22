#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CR Tracker App (from WhatsApp "Export Chat" .txt)
===================================================
A small desktop GUI app that reads the plain-text file produced by
WhatsApp's own "Export Chat" feature (phone: chat > ... > More > Export
chat > Without Media), lets you pick a date range, and builds/updates a
"CR Tracker" Excel file from the Sign-in / Sign-out messages found in
that range — same tracker format as the other tools in this set.

Usage:
    pip install openpyxl
    python cr_tracker_from_txt_app.py

To make a clickable .exe (optional, on a Windows machine):
    pip install pyinstaller openpyxl
    pyinstaller --onefile --windowed --name "CR Tracker (from chat export)" cr_tracker_from_txt_app.py
"""

import re
import shutil
import datetime
import threading
from pathlib import Path

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ======================================================================
# 1) Parsing the raw WhatsApp "Export Chat" .txt file
# ======================================================================

# Matches the start of a new message line, e.g.:
#   "14/12/2023, 16:46 - Mhmd AminFo: *MS-FO Duty*"
#   "11/12/2023, 16:40 - Me was added"          (system message, no ": ")
LINE_RE = re.compile(
    r"^(\d{1,2})/(\d{1,2})/(\d{2,4}),\s*(\d{1,2}):(\d{2})(?::(\d{2}))?\s*[\u202f\u200e]?(AM|PM|am|pm)?\s*-\s*(.*)$"
)


def load_messages_from_txt(txt_path, date_order="DMY"):
    """
    Reads a WhatsApp Export Chat .txt file and returns a list of
    {"time": datetime, "sender": str, "body": str} dicts, with
    multi-line messages correctly joined together.
    """
    text = Path(txt_path).read_text(encoding="utf-8", errors="ignore")
    # Some exports use a left-to-right mark (U+200E) at the start of each line
    text = text.replace("\u200e", "")

    lines = text.splitlines()
    messages = []
    current = None

    for line in lines:
        m = LINE_RE.match(line)
        if m:
            if current:
                messages.append(current)
            d1, d2, year, hh, mm, ss, ampm, rest = m.groups()
            d1, d2 = int(d1), int(d2)
            if date_order == "MDY":
                month, day = d1, d2
            else:
                day, month = d1, d2
            year = int(year)
            if year < 100:
                year += 2000
            hh = int(hh)
            mm = int(mm)
            ss = int(ss) if ss else 0
            if ampm:
                is_pm = ampm.lower() == "pm"
                if hh == 12:
                    hh = 12 if is_pm else 0
                elif is_pm:
                    hh += 12
            try:
                ts = datetime.datetime(year, month, day, hh, mm, ss)
            except ValueError:
                current = None
                continue

            if ": " in rest:
                sender, body = rest.split(": ", 1)
            else:
                sender, body = "", rest

            current = {"time": ts, "sender": sender.strip(), "body": body}
        else:
            if current is not None:
                current["body"] += "\n" + line

    if current:
        messages.append(current)

    messages.sort(key=lambda m: m["time"])
    return messages


def scan_date_range(txt_path, date_order="DMY"):
    """Quick helper to find the earliest/latest message date in the file,
    used to pre-fill the From/To fields."""
    messages = load_messages_from_txt(txt_path, date_order)
    if not messages:
        return None, None
    return messages[0]["time"].date(), messages[-1]["time"].date()


# ======================================================================
# 2) CR message parsing (same logic as the other tools in this set)
# ======================================================================

def _clean(text):
    text = text.replace("\\*", "*").replace("*", "")
    text = text.replace("\\n", "\n").replace("\\t", "\t")
    return text


def _find_field(text, *labels):
    for label in labels:
        pattern = re.escape(label).replace(r"\ ", r"\s*")
        m = re.search(pattern + r"\s*[:\-]?\s*(.+)", text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            val = re.split(r"[\r\n]", val)[0].strip(" :-\t")
            if val:
                return val
    return None


CR_NUMBER_RE = re.compile(r"CR-\d{6,10}-\d{4,10}", re.IGNORECASE)


def parse_cr_message(raw_body, msg_time):
    text = _clean(raw_body)
    lower = text.lower()

    cr_match = CR_NUMBER_RE.search(text)
    if not cr_match:
        return None

    if re.search(r"sign[\s_\-]*out", lower):
        direction = "out"
    elif re.search(r"sign[\s_\-]*in", lower):
        direction = "in"
    else:
        return None

    site_id_raw = _find_field(text, "SiteID", "Site ID", "Site Id")
    if not site_id_raw:
        return None
    if site_id_raw.startswith("(") and "__" in site_id_raw:
        site_id = site_id_raw
    else:
        m_primary = re.match(r"[A-Za-z0-9]+", site_id_raw)
        site_id = m_primary.group(0) if m_primary else site_id_raw

    type_of_cr_raw = _find_field(text, "Type of CR") or ""
    type_of_cr = normalize_type_of_cr(type_of_cr_raw)

    return {
        "direction": direction,
        "cr_number": cr_match.group(0).upper(),
        "site_id": site_id,
        "type_of_cr": type_of_cr,
        "time": msg_time,
    }


def normalize_type_of_cr(raw):
    r = raw.lower()
    if "pre" in r:
        return "Preapprove"
    if "approv" in r:
        return "Approved"
    if "reject" in r:
        return "Rejected"
    return raw.strip()


FOR_SITE_RE = re.compile(r"for\s+sites?\s*[:\-]?\s*(.+)", re.IGNORECASE)


def build_site_name_map(messages):
    site_map = {}
    for m in messages:
        text = _clean(m["body"])
        line_match = FOR_SITE_RE.search(text)
        if not line_match:
            continue
        rest = re.split(r"[\r\n]", line_match.group(1))[0].strip()
        if rest.count("(") or rest.count(")"):
            continue
        parts = re.split(r"\s*[_\-]\s*", rest)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) == 2:
            site_id, site_name = parts
            if re.match(r"^[A-Za-z0-9]{4,}$", site_id):
                site_map[site_id.upper()] = site_name
    return site_map


def lookup_site_name(site_id, site_map):
    key = site_id.strip().upper()
    if key in site_map:
        return site_map[key]
    first_token = re.split(r"[_(]", key)[0]
    return site_map.get(first_token, "")


def build_tracker_rows(messages, site_map):
    rows = {}
    order = []

    for m in messages:
        parsed = parse_cr_message(m["body"], m["time"])
        if not parsed:
            continue

        key = (parsed["cr_number"], parsed["site_id"])
        if key not in rows:
            rows[key] = {
                "CR Number": parsed["cr_number"],
                "Type of CR": parsed["type_of_cr"],
                "Main site Name": lookup_site_name(parsed["site_id"], site_map),
                "Main site ID": parsed["site_id"],
                "ST": None,
                "ET": None,
                "Sign in time": None,
                "Sign out time": None,
            }
            order.append(key)

        row = rows[key]
        if not row["Type of CR"] and parsed["type_of_cr"]:
            row["Type of CR"] = parsed["type_of_cr"]
        if not row["Main site Name"]:
            row["Main site Name"] = lookup_site_name(parsed["site_id"], site_map)

        t = parsed["time"].time()
        if parsed["direction"] == "in" and row["Sign in time"] is None:
            row["Sign in time"] = t
        elif parsed["direction"] == "out":
            row["Sign out time"] = t

    return [rows[k] for k in order]


# ======================================================================
# 3) Writing / merging the tracker Excel (same as the other tools)
# ======================================================================

COLUMNS = ["CR Number", "Type of CR", "Main site Name", "Main site ID",
           "ST", "ET", "Sign in time", "Sign out time"]

HEADER_FILL = PatternFill("solid", fgColor="00B0F0")
HEADER_FONT = Font(name="Calibri", size=14, bold=True, italic=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=11)


def load_existing_rows(tracker_path):
    existing = {}
    order = []
    if not tracker_path.exists():
        return existing, order

    wb = openpyxl.load_workbook(tracker_path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        def g(col):
            i = idx.get(col)
            return row[i] if i is not None and i < len(row) else None

        cr = g("CR Number")
        site = g("Main site ID")
        if not cr or not site:
            continue
        key = (str(cr).strip().upper(), str(site).strip())
        existing[key] = {
            "CR Number": cr,
            "Type of CR": g("Type of CR"),
            "Main site Name": g("Main site Name"),
            "Main site ID": site,
            "ST": g("ST"),
            "ET": g("ET"),
            "Sign in time": g("Sign in time"),
            "Sign out time": g("Sign out time"),
        }
        order.append(key)
    return existing, order


def merge_rows(existing, existing_order, new_rows):
    for row in new_rows:
        key = (row["CR Number"].strip().upper(), row["Main site ID"].strip())
        if key not in existing:
            existing[key] = row
            existing_order.append(key)
        else:
            cur = existing[key]
            for field in ("Type of CR", "Main site Name", "ST", "ET"):
                if not cur.get(field) and row.get(field):
                    cur[field] = row[field]
            if not cur.get("Sign in time") and row.get("Sign in time"):
                cur["Sign in time"] = row["Sign in time"]
            if row.get("Sign out time"):
                cur["Sign out time"] = row["Sign out time"]
    return existing, existing_order


def write_tracker(all_rows_dict, order, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for c, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, key in enumerate(order, start=2):
        row = all_rows_dict[key]
        for c, col_name in enumerate(COLUMNS, start=1):
            val = row.get(col_name)
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            if col_name in ("ST", "ET") and val is not None:
                cell.number_format = "m/d/yy h:mm"
            elif col_name == "Sign in time" and val is not None:
                cell.number_format = "h:mm"
            elif col_name == "Sign out time" and val is not None:
                cell.number_format = "h:mm:ss AM/PM"

    widths = [24, 14, 18, 16, 16, 16, 13, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(output_path)


# ======================================================================
# 4) Main pipeline
# ======================================================================

def process(txt_path, tracker_path, date_from, date_to, date_order, log):
    all_messages = load_messages_from_txt(txt_path, date_order)
    log(f"Total messages in the chat file: {len(all_messages)}")

    site_map = build_site_name_map(all_messages)
    log(f"Site IDs matched to a name (from the whole file): {len(site_map)}")

    if date_from:
        start = datetime.datetime.combine(date_from, datetime.time.min)
    else:
        start = None
    if date_to:
        end = datetime.datetime.combine(date_to, datetime.time.max)
    else:
        end = None

    ranged = [
        m for m in all_messages
        if (start is None or m["time"] >= start) and (end is None or m["time"] <= end)
    ]
    log(f"Messages within the selected date range: {len(ranged)}")

    new_rows = build_tracker_rows(ranged, site_map)
    log(f"Sign-in/out messages understood in this range: {len(new_rows)}")

    # Always build a brand-new tracker from this range only — no merging
    # with whatever might already exist at the destination path.
    order = [(row["CR Number"].strip().upper(), row["Main site ID"].strip()) for row in new_rows]
    all_rows_dict = {key: row for key, row in zip(order, new_rows)}
    write_tracker(all_rows_dict, order, tracker_path)

    log(f"Saved a new tracker with {len(order)} rows to: {tracker_path}")

    unmapped = sorted({r['Main site ID'] for r in new_rows if not r['Main site Name']})
    if unmapped:
        log(f"Sites with no known name ({len(unmapped)}): {', '.join(unmapped)}")

    return len(order)


# ======================================================================
# 5) GUI
# ======================================================================

class CRTrackerTxtApp:
    def __init__(self, root):
        self.root = root
        root.title("CR Tracker – from WhatsApp Chat Export (.txt)")
        root.geometry("680x560")

        pad = {"padx": 10, "pady": 6}
        frm = ttk.Frame(root)
        frm.pack(fill="x", **pad)

        ttk.Label(frm, text="WhatsApp chat export file (.txt):").grid(row=0, column=0, sticky="w")
        self.txt_var = tk.StringVar()
        ttk.Entry(frm, textvariable=self.txt_var, width=55).grid(row=1, column=0, sticky="we")
        ttk.Button(frm, text="Browse...", command=self.pick_txt).grid(row=1, column=1, padx=6)

        ttk.Label(frm, text="Date format used in the file:").grid(row=2, column=0, sticky="w", pady=(12, 0))
        self.date_order_var = tk.StringVar(value="DMY")
        order_row = ttk.Frame(frm)
        order_row.grid(row=3, column=0, sticky="w")
        ttk.Radiobutton(order_row, text="DD/MM/YYYY (most phones)", variable=self.date_order_var, value="DMY").pack(side="left")
        ttk.Radiobutton(order_row, text="MM/DD/YYYY (US phones)", variable=self.date_order_var, value="MDY").pack(side="left", padx=10)

        date_row = ttk.Frame(frm)
        date_row.grid(row=4, column=0, columnspan=2, sticky="we", pady=(12, 0))
        ttk.Label(date_row, text="From (YYYY-MM-DD):").pack(side="left")
        self.from_var = tk.StringVar()
        ttk.Entry(date_row, textvariable=self.from_var, width=12).pack(side="left", padx=(4, 16))
        ttk.Label(date_row, text="To (YYYY-MM-DD):").pack(side="left")
        self.to_var = tk.StringVar()
        ttk.Entry(date_row, textvariable=self.to_var, width=12).pack(side="left", padx=4)

        ttk.Label(frm, text="Save the new CR Tracker as:").grid(
            row=5, column=0, sticky="w", pady=(14, 0)
        )
        self.tracker_var = tk.StringVar()
        ttk.Entry(frm, textvariable=self.tracker_var, width=55).grid(row=6, column=0, sticky="we")
        ttk.Button(frm, text="Browse...", command=self.pick_tracker).grid(row=6, column=1, padx=6)

        frm.columnconfigure(0, weight=1)

        btn_row = ttk.Frame(root)
        btn_row.pack(pady=12)

        self.run_btn = ttk.Button(btn_row, text="▶  Run", command=self.run_clicked)
        self.run_btn.grid(row=0, column=0, padx=6)

        self.download_btn = ttk.Button(
            btn_row, text="⬇  Download CR Tracker", command=self.download_clicked, state="disabled"
        )
        self.download_btn.grid(row=0, column=1, padx=6)

        self.log_box = scrolledtext.ScrolledText(root, height=16, wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=10, pady=6)
        self.log_box.configure(state="disabled")

    def pick_txt(self):
        path = filedialog.askopenfilename(
            title="Choose the WhatsApp chat export .txt file",
            filetypes=[("Text files", "*.txt")]
        )
        if not path:
            return
        self.txt_var.set(path)
        if not self.tracker_var.get():
            self.tracker_var.set(str(Path(path).with_name("CR_Tracker.xlsx")))

        def scan():
            try:
                lo, hi = scan_date_range(path, self.date_order_var.get())
                if lo and hi:
                    self.from_var.set(lo.isoformat())
                    self.to_var.set(hi.isoformat())
                    self.log(f"File scanned. Messages range from {lo} to {hi}.")
            except Exception as e:
                self.log(f"Could not pre-scan dates: {e}")

        threading.Thread(target=scan, daemon=True).start()

    def pick_tracker(self):
        path = filedialog.asksaveasfilename(
            title="Choose / name the CR Tracker file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.tracker_var.set(path)

    def log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.root.update_idletasks()

    def run_clicked(self):
        txt = self.txt_var.get().strip()
        tracker = self.tracker_var.get().strip()

        if not txt or not Path(txt).exists():
            messagebox.showerror("Error", "Please choose a valid WhatsApp chat export .txt file.")
            return
        if not tracker:
            messagebox.showerror("Error", "Please choose a location/name for the CR Tracker file.")
            return

        date_from = date_to = None
        try:
            if self.from_var.get().strip():
                date_from = datetime.date.fromisoformat(self.from_var.get().strip())
            if self.to_var.get().strip():
                date_to = datetime.date.fromisoformat(self.to_var.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Dates must be in YYYY-MM-DD format.")
            return

        self.run_btn.configure(state="disabled")
        self.download_btn.configure(state="disabled")
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

        def worker():
            try:
                process(Path(txt), Path(tracker), date_from, date_to,
                        self.date_order_var.get(), self.log)
                self.download_btn.configure(state="normal")
                messagebox.showinfo("Done", "The CR Tracker file was updated successfully.")
            except Exception as e:
                self.log(f"❌ Error: {e}")
                messagebox.showerror("Error", str(e))
            finally:
                self.run_btn.configure(state="normal")

        threading.Thread(target=worker, daemon=True).start()

    def download_clicked(self):
        tracker = self.tracker_var.get().strip()
        if not tracker or not Path(tracker).exists():
            messagebox.showerror("Error", "No CR Tracker file to download yet. Click Run first.")
            return

        dest = filedialog.asksaveasfilename(
            title="Save CR Tracker as...",
            defaultextension=".xlsx",
            initialfile=Path(tracker).name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not dest:
            return

        try:
            shutil.copyfile(tracker, dest)
            self.log(f"Downloaded a copy to: {dest}")
            messagebox.showinfo("Downloaded", f"Saved a copy to:\n{dest}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save the file: {e}")


def main():
    root = tk.Tk()
    CRTrackerTxtApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
